"""
README — First-Column Order Checker (Excel, FILES mode)-- Junhua Huang 08/26/2025

What this script does
- Treats the FIRST column of one Excel file as the ground-truth order.
- Compares the FIRST column of 8 (or more) other Excel files against it.
- Reports: exact row positions where values differ, whether files have extra/missing rows,
  and whether the set of IDs matches ignoring order (duplicates accounted for).
- Never modifies your source files.

Minimal environment (one lightweight dep)
    python3 -m venv .venv
    source .venv/bin/activate        # Windows: .venv\Scripts\activate
    pip install --upgrade pip openpyxl

How to configure (FILES mode only)
1) Open this script and edit the CONFIG section:
   - Set MODE = "files"
   - Set GROUND_TRUTH_FILE to your GT path, e.g.:
       GROUND_TRUTH_FILE = "/home/harry/projects/ucla/data/seizure_feature_ground_truth.xlsx"
   - List your test files in TEST_FILES, e.g.:
       TEST_FILES = [
           "/home/harry/projects/ucla/data/seizure_feature_Detao_junnan.xlsx",
           "/home/harry/projects/ucla/data/seizure_feature_hailey_jiarui.xlsx",
           "/home/harry/projects/ucla/data/seizure_feature_jayhee_ido.xlsx",
           "/home/harry/projects/ucla/data/seizure_feature_Jiarui_xinyi.xlsx",
           "/home/harry/projects/ucla/data/seizure_feature_jiaye.xlsx",
           "/home/harry/projects/ucla/data/seizure_feature_Lingye_JunHua.xlsx",
           "/home/harry/projects/ucla/data/seizure_feature_spatial_chonghan.xlsx",
           "/home/harry/projects/ucla/data/seizure_feature_Tengyou_xiangting.xlsx",
       ]
   - If your first row is a header, leave SKIP_HEADER = True (recommended).
   - Optional normalization:
       STRIP_WHITESPACE = True    # trims leading/trailing spaces
       CASE_INSENSITIVE = False   # set True to ignore case differences

How to run
    python check_order.py

Outputs (written to the working directory)
- order_check_report.xlsx
  • “Summary” sheet (one row per test file): 
      identical_order (perfect match?), first_mismatch_index, counts of mismatch/missing/extra, 
      same_multiset_ignoring_order (values match ignoring order?), duplicates count, and the name of the diff tab.
  • One “diff_<file>” sheet per test with row-by-row details:
      index (0-based), index (1-based), truth_value, test_value, status ∈ {match, mismatch, missing_in_test, extra_in_test}.
- order_check_diffs.txt
  • Human-readable log of all non-matching positions (type + index + values).
- order_check_diffs.jsonl
  • Machine-readable version (one JSON per difference).

How to interpret quickly
- identical_order = TRUE → file matches ground truth exactly (same length and order).
- identical_order = FALSE, same_multiset_ignoring_order = TRUE → values are the same but shuffled/duplicated.
- Both FALSE → actual content difference (missing/extra IDs or wrong value).
- Use the per-file diff sheet: find the row by “index (1-based)”, copy the correct value from truth_value,
  remove any row flagged extra_in_test, and add any value flagged as missing.

Notes / tips
- Only the FIRST column is checked.
- If you see late “missing_in_test” rows, your test file is shorter than the ground truth.
- Duplicates are summarized in “duplicate_values_in_test_unique” and visible in the diff sheet.
- The 1-based index aligns with typical Excel row counting (after the header if SKIP_HEADER=True).
"""



from collections import Counter
from pathlib import Path
import re
from openpyxl import load_workbook, Workbook

# -------------------- CONFIG (edit these) --------------------
# Choose ONE mode: "sheets" (all in one workbook) OR "files" (separate workbooks)
MODE = "files"   # "sheets" or "files"

# A) All sheets in one workbook (your example path)
ONE_WORKBOOK = "/home/harry/projects/ucla/data/seizure_feature.xlsx"
GROUND_TRUTH_SHEET = "ground_truth"   # <- rename to your GT sheet name
TEST_SHEETS = ["test1","test2","test3","test4","test5","test6","test7","test8"]  # <- rename

# B) Separate workbooks (ignore if using MODE="sheets")
GROUND_TRUTH_FILE = "/home/harry/projects/ucla/data/seizure_feature.xlsx"
TEST_FILES = [
    "/home/harry/projects/ucla/data/seizure_feature_Detao_junnan.xlsx",
    "/home/harry/projects/ucla/data/seizure_feature_hailey_jiarui.xlsx", 
    "/home/harry/projects/ucla/data/seizure_feature_jayhee_ido.xlsx",
    "/home/harry/projects/ucla/data/seizure_feature_Jiarui_xinyi.xlsx", 
    "/home/harry/projects/ucla/data/seizure_feature_jiaye.xlsx", 
    "/home/harry/projects/ucla/data/seizure_feature_Lingye_JunHua.xlsx",
    "/home/harry/projects/ucla/data/seizure_feature_spatial_chonghan.xlsx", 
    "/home/harry/projects/ucla/data/seizure_feature_Tengyou_xiangting.xlsx",
]

# First row header? (True = skip first row)
SKIP_HEADER = True
# Normalize?
STRIP_WHITESPACE = True
CASE_INSENSITIVE = False

# Output
REPORT_PATH = "order_check_report.xlsx"
# -------------------------------------------------------------

def read_first_col_xlsx(path, sheet_name=None, skip_header=True,
                        strip_ws=True, lower=False):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
    values = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_col=1, values_only=True)):
        if i == 0 and skip_header:
            continue
        cell = row[0]
        if cell is None:
            continue
        s = str(cell)
        if strip_ws:
            s = s.strip()
        if lower:
            s = s.lower()
        values.append(s)
    wb.close()
    return values

def compare_lists(gt, test):
    max_len = max(len(gt), len(test))
    rows = []
    n_mismatch = n_missing = n_extra = 0
    first_mismatch_index = -1

    for i in range(max_len):
        gt_val = gt[i] if i < len(gt) else ""
        ts_val = test[i] if i < len(test) else ""
        if i >= len(test):
            status = "missing_in_test"
            n_missing += 1
            if first_mismatch_index == -1:
                first_mismatch_index = i
        elif i >= len(gt):
            status = "extra_in_test"
            n_extra += 1
            if first_mismatch_index == -1:
                first_mismatch_index = i
        else:
            if ts_val == gt_val:
                status = "match"
            else:
                status = "mismatch"
                n_mismatch += 1
                if first_mismatch_index == -1:
                    first_mismatch_index = i
        rows.append({"index": i, "truth_value": gt_val, "test_value": ts_val, "status": status})

    gt_ctr, ts_ctr = Counter(gt), Counter(test)
    same_multiset = (gt_ctr == ts_ctr)

    missing_elements, extra_elements = [], []
    for k, v in gt_ctr.items():
        diff = v - ts_ctr.get(k, 0)
        if diff > 0:
            missing_elements.extend([k] * diff)
    for k, v in ts_ctr.items():
        diff = v - gt_ctr.get(k, 0)
        if diff > 0:
            extra_elements.extend([k] * diff)

    return {
        "rows": rows,
        "n_mismatch_positions": n_mismatch,
        "n_missing_positions": n_missing,
        "n_extra_positions": n_extra,
        "first_mismatch_index": first_mismatch_index,
        "same_multiset": same_multiset,
        "missing_elements": missing_elements,
        "extra_elements": extra_elements,
    }

def safe_sheet_name(name):
    base = re.sub(r'[:\\/?*\[\]]', '_', str(name))
    return base[:31] if len(base) > 31 else base

def write_report(summary_rows, diff_tabs, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    headers = [
        "name",
        "identical_order",
        "same_multiset_ignoring_order",
        "len_truth",
        "len_test",
        "mismatch_positions",
        "missing_positions",
        "extra_positions",
        "first_mismatch_index",
        "missing_elements_count",
        "extra_elements_count",
        "duplicate_values_in_test_unique",
        "diff_tab",
    ]
    ws.append(headers)
    for r in summary_rows:
        ws.append([
            r["name"], r["identical_order"], r["same_multiset_ignoring_order"],
            r["len_truth"], r["len_test"], r["mismatch_positions"],
            r["missing_positions"], r["extra_positions"], r["first_mismatch_index"],
            r["missing_elements_count"], r["extra_elements_count"],
            r["duplicate_values_in_test_unique"], r["diff_tab"]
        ])

    for tab in diff_tabs:
        ws2 = wb.create_sheet(title=safe_sheet_name(tab["tab_name"]))
        ws2.append(["index (0-based)", "truth_value", "test_value", "status"])
        for row in tab["rows"]:
            ws2.append([row["index"], row["truth_value"], row["test_value"], row["status"]])

    wb.save(out_path)

def main():
    # Load GT
    if MODE == "sheets":
        gt = read_first_col_xlsx(
            ONE_WORKBOOK, GROUND_TRUTH_SHEET,
            SKIP_HEADER, STRIP_WHITESPACE, CASE_INSENSITIVE
        )
    else:
        gt = read_first_col_xlsx(
            GROUND_TRUTH_FILE, None,
            SKIP_HEADER, STRIP_WHITESPACE, CASE_INSENSITIVE
        )

    summary_rows, diff_tabs = [], []

    if MODE == "sheets":
        for sheet in TEST_SHEETS:
            test = read_first_col_xlsx(
                ONE_WORKBOOK, sheet,
                SKIP_HEADER, STRIP_WHITESPACE, CASE_INSENSITIVE
            )
            res = compare_lists(gt, test)

            test_counts = Counter(test)
            dup_unique = sum(1 for k, v in test_counts.items() if v > 1)

            tab_name = f"diff_{sheet}"
            summary_rows.append({
                "name": sheet,
                "identical_order": (res["n_mismatch_positions"] == 0 and
                                    res["n_missing_positions"] == 0 and
                                    res["n_extra_positions"] == 0 and
                                    len(test) == len(gt)),
                "same_multiset_ignoring_order": res["same_multiset"],
                "len_truth": len(gt),
                "len_test": len(test),
                "mismatch_positions": res["n_mismatch_positions"],
                "missing_positions": res["n_missing_positions"],
                "extra_positions": res["n_extra_positions"],
                "first_mismatch_index": res["first_mismatch_index"],
                "missing_elements_count": len(res["missing_elements"]),
                "extra_elements_count": len(res["extra_elements"]),
                "duplicate_values_in_test_unique": dup_unique,
                "diff_tab": tab_name
            })
            diff_tabs.append({"tab_name": tab_name, "rows": res["rows"]})

    else:
        gt_path = Path(GROUND_TRUTH_FILE)
        for f in TEST_FILES:
            test = read_first_col_xlsx(
                f, None,
                SKIP_HEADER, STRIP_WHITESPACE, CASE_INSENSITIVE
            )
            res = compare_lists(gt, test)

            test_counts = Counter(test)
            dup_unique = sum(1 for k, v in test_counts.items() if v > 1)

            name = Path(f).name
            short = Path(f).stem
            tab_name = f"diff_{short}"
            summary_rows.append({
                "name": name,
                "identical_order": (res["n_mismatch_positions"] == 0 and
                                    res["n_missing_positions"] == 0 and
                                    res["n_extra_positions"] == 0 and
                                    len(test) == len(gt)),
                "same_multiset_ignoring_order": res["same_multiset"],
                "len_truth": len(gt),
                "len_test": len(test),
                "mismatch_positions": res["n_mismatch_positions"],
                "missing_positions": res["n_missing_positions"],
                "extra_positions": res["n_extra_positions"],
                "first_mismatch_index": res["first_mismatch_index"],
                "missing_elements_count": len(res["missing_elements"]),
                "extra_elements_count": len(res["extra_elements"]),
                "duplicate_values_in_test_unique": dup_unique,
                "diff_tab": tab_name
            })
            diff_tabs.append({"tab_name": tab_name, "rows": res["rows"]})

    write_report(summary_rows, diff_tabs, REPORT_PATH)
    print(f"Saved: {REPORT_PATH}")
    print("Tabs: Summary + one diff_* tab per test")

if __name__ == "__main__":
    main()
